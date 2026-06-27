import os
import re
import json
import secrets
import random
import string
import threading
import requests
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, request, jsonify, send_file, session, redirect, url_for, render_template
from flask_cors import CORS
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import psycopg2
from psycopg2.extras import RealDictCursor
from psycopg2.pool import SimpleConnectionPool
import smtplib
import ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import logging
from logging.handlers import RotatingFileHandler
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

app = Flask(__name__)
CORS(app)

# ============================================================================
# CONFIGURATION
# ============================================================================

app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

# Database
DATABASE_URL = os.environ.get('DATABASE_URL', 'postgresql://postgres:password@localhost:5432/parser_db')

# Email Configuration (for OTP)
MAIL_SERVER = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
MAIL_PORT = int(os.environ.get('MAIL_PORT', 465))
MAIL_USE_TLS = os.environ.get('MAIL_USE_TLS', 'False').lower() == 'true'
MAIL_USERNAME = os.environ.get('MAIL_USERNAME', '')
MAIL_PASSWORD = os.environ.get('MAIL_PASSWORD', '')
MAIL_FROM = os.environ.get('MAIL_FROM', MAIL_USERNAME)

# OTP Settings
OTP_EXPIRY_MINUTES = 10
OTP_LENGTH = 6

# Upload settings
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf'}
MAX_FILE_SIZE = 50 * 1024 * 1024
SESSION_TIMEOUT = 3600

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ============================================================================
# DATABASE CONNECTION POOL
# ============================================================================

db_pool = None

def get_db_pool():
    global db_pool
    if db_pool is None:
        try:
            db_pool = SimpleConnectionPool(
                minconn=1,
                maxconn=20,
                dsn=DATABASE_URL
            )
            print("✓ Database connection pool created")
        except Exception as e:
            print(f"✗ Database connection error: {e}")
            raise
    return db_pool

def get_db_connection():
    pool = get_db_pool()
    return pool.getconn()

def return_db_connection(conn):
    pool = get_db_pool()
    pool.putconn(conn)

# ============================================================================
# DATABASE INITIALIZATION
# ============================================================================

def init_db():
    """Initialize PostgreSQL database with required tables"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Create users table with email_verified field
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username VARCHAR(50) UNIQUE NOT NULL,
                email VARCHAR(100) UNIQUE NOT NULL,
                password_hash VARCHAR(255) NOT NULL,
                email_verified BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                last_login TIMESTAMP,
                is_active BOOLEAN DEFAULT TRUE
            )
        ''')
        
        # Create email_verifications table for OTP
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS email_verifications (
                id SERIAL PRIMARY KEY,
                email VARCHAR(100) NOT NULL,
                otp VARCHAR(6) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                expires_at TIMESTAMP NOT NULL,
                is_used BOOLEAN DEFAULT FALSE
            )
        ''')
        
        # Create invoices table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS invoices (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                invoice_number TEXT NOT NULL,
                invoice_date TEXT NOT NULL,
                terms_of_payment TEXT,
                terms_of_delivery TEXT,
                currency TEXT,
                amount TEXT,
                submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Create shipments table with invoice_number column
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS shipments (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                ship_bill_no TEXT NOT NULL,
                ship_billing_date TEXT NOT NULL,
                invoice_number TEXT,
                submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Create indexes
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_invoices_user_id ON invoices(user_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_invoices_submitted_at ON invoices(submitted_at)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_shipments_user_id ON shipments(user_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_shipments_submitted_at ON shipments(submitted_at)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_shipments_invoice_number ON shipments(invoice_number)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_email_verifications_email ON email_verifications(email)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_email_verifications_otp ON email_verifications(otp)')
        
        conn.commit()
        print("✓ Database initialized successfully")
        
    except Exception as e:
        conn.rollback()
        print(f"✗ Database initialization error: {e}")
        raise
    finally:
        cursor.close()
        return_db_connection(conn)

# ============================================================================
# AUTHENTICATION DECORATOR
# ============================================================================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'status': 'error', 'message': 'Please login first', 'redirect': '/login'}), 401
        return f(*args, **kwargs)
    return decorated_function

def get_current_user():
    if 'user_id' in session:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        try:
            cursor.execute('SELECT id, username, email, email_verified FROM users WHERE id = %s', (session['user_id'],))
            user = cursor.fetchone()
            return user
        finally:
            cursor.close()
            return_db_connection(conn)
    return None

def get_user_id():
    return session.get('user_id')

# ============================================================================
# EMAIL HELPER FUNCTIONS (NON-BLOCKING WITH DEBUG)
# ============================================================================

def generate_otp():
    """Generate a 6-digit OTP"""
    return ''.join(random.choices(string.digits, k=OTP_LENGTH))


def send_otp_email_async(email, otp, username):
    """Send OTP via SendGrid Web API"""
    
    print(f"🔍 ===== SENDGRID WEB API =====")
    print(f"🔍 Sending to: {email}")
    print(f"🔍 OTP: {otp}")
    print(f"🔍 ============================")
    
    if not MAIL_PASSWORD:
        print(f"⚠ SendGrid API key not configured. OTP for {email}: {otp}")
        return True
    
    try:
        # SendGrid Web API URL
        url = "https://api.sendgrid.com/v3/mail/send"
        
        # Email content
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; background-color: #f4f4f4; padding: 20px; }}
                .container {{ max-width: 500px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                .header {{ text-align: center; border-bottom: 2px solid #1b5e20; padding-bottom: 20px; }}
                .header h1 {{ color: #1b5e20; margin: 0; }}
                .otp-code {{ font-size: 32px; font-weight: bold; color: #1b5e20; text-align: center; padding: 20px; background: #e8f5e9; border-radius: 8px; margin: 20px 0; letter-spacing: 8px; }}
                .footer {{ text-align: center; color: #999; font-size: 12px; margin-top: 20px; }}
                .expiry {{ color: #666; font-size: 14px; text-align: center; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>📄 Invoice Parser Portal</h1>
                </div>
                <p>Hello <strong>{username}</strong>,</p>
                <p>Thank you for registering! Please verify your email address by entering the OTP below:</p>
                <div class="otp-code">{otp}</div>
                <p class="expiry">This OTP expires in <strong>10 minutes</strong></p>
                <p>If you didn't request this, please ignore this email.</p>
                <div class="footer">
                    <p>Invoice & Shipment Bill Parser Portal</p>
                    <p>This is an automated message, please do not reply.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Email data
        data = {
            "personalizations": [
                {
                    "to": [{"email": email}],
                    "subject": "Verify Your Email - Invoice Parser Portal"
                }
            ],
            "from": {"email": MAIL_FROM or MAIL_USERNAME},
            "content": [
                {
                    "type": "text/html",
                    "value": html_content
                }
            ]
        }
        
        # Headers
        headers = {
            "Authorization": f"Bearer {MAIL_PASSWORD}",
            "Content-Type": "application/json"
        }
        
        print(f"🔍 Sending request to SendGrid API...")
        
        response = requests.post(url, json=data, headers=headers, timeout=30)
        
        if response.status_code == 202:
            print(f"✓ Email sent successfully to {email}")
            return True
        else:
            print(f"✗ SendGrid API error: {response.status_code} - {response.text}")
            return False
            
    except requests.exceptions.Timeout:
        print(f"✗ SendGrid API timeout")
        return False
    except Exception as e:
        print(f"✗ Email send error: {str(e)}")
        return False
    
def send_otp_email(email, otp, username):
    """Send OTP via email (non-blocking)"""
    # Always print OTP to console (for Railway logs)
    print(f"📧 OTP for {email}: {otp}")
    
    # Send email in background thread (non-blocking)
    thread = threading.Thread(target=send_otp_email_async, args=(email, otp, username))
    thread.daemon = True
    thread.start()
    
    return True

def send_password_reset_email(email, otp, username):
    """Send password reset OTP via email (non-blocking)"""
    print(f"📧 Password Reset OTP for {email}: {otp}")
    
    # Send in background thread
    thread = threading.Thread(target=send_otp_email_async, args=(email, otp, username))
    thread.daemon = True
    thread.start()
    
    return True

# ============================================================================
# HELPER FUNCTIONS - PDF EXTRACTION
# ============================================================================

def allowed_file(filename):
    """Check if file has allowed extension"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def convert_date_format(date_str):
    """Convert various date formats to DD-MM-YYYY format"""
    if not date_str:
        return ""
    
    date_str = date_str.strip()
    
    # Handle DD-MMM-YY format
    month_map = {
        'JAN': '01', 'FEB': '02', 'MAR': '03', 'APR': '04', 
        'MAY': '05', 'JUN': '06', 'JUL': '07', 'AUG': '08',
        'SEP': '09', 'OCT': '10', 'NOV': '11', 'DEC': '12'
    }
    
    # Try DD-MMM-YY format
    match = re.search(r'(\d{1,2})\s*[-/]?\s*([A-Z]{3})\s*[-/]?\s*(\d{2,4})', date_str, re.IGNORECASE)
    if match:
        day, month_str, year = match.groups()
        month_str_upper = month_str.upper()
        if month_str_upper in month_map:
            month = month_map[month_str_upper]
            year = int(year)
            if year < 100:
                year = 2000 + year if year < 50 else 1900 + year
            return f"{int(day):02d}-{month}-{year}"
    
    # Try DD.MM.YYYY or DD/MM/YYYY format
    match = re.search(r'(\d{1,2})\s*[./]\s*(\d{1,2})\s*[./]\s*(\d{4})', date_str)
    if match:
        day, month, year = match.groups()
        return f"{int(day):02d}-{int(month):02d}-{year}"
    
    # Try DD-MM-YYYY format
    match = re.search(r'(\d{1,2})\s*-\s*(\d{1,2})\s*-\s*(\d{4})', date_str)
    if match:
        day, month, year = match.groups()
        return f"{int(day):02d}-{int(month):02d}-{year}"
    
    return date_str

def extract_invoice_data(pdf_path):
    """Extract invoice data from PDF using pdfplumber"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"

            result = {
                'invoiceNumber': '',
                'invoiceDate': '',
                'termsOfPayment': '',
                'termsOfDelivery': '',
                'currency': '',
                'amount': ''
            }
            
            # Extract Invoice Number
            patterns = [
                r'Invoice\s*No\s*&\s*Date\s*\n.*?(\d{10})',
                r'Invoice\s*No\s*&\s*Date\s*[^\d]*(\d+)',
                r'Invoice\s*No[.\s:]+(\d+)',
                r'INV[OICE]*\s*No[.\s:]+(\d+)',
            ]

            for pattern in patterns:
                match = re.search(pattern, full_text, re.IGNORECASE | re.DOTALL)
                if match:
                    result['invoiceNumber'] = match.group(1).strip()
                    break
            
            # Extract Invoice Date
            patterns = [
                r'Invoice\s*(?:No\s*&\s*)?Date[.\s:]*([^\n]+?)(?:\n|$)',
                r'Dt:\s*(\d{1,2}[./\-]\d{1,2}[./\-]\d{2,4})',
                r'Date[.\s:]*(\d{1,2}[./\-]\d{1,2}[./\-]\d{2,4})'
            ]
            for pattern in patterns:
                match = re.search(pattern, full_text, re.IGNORECASE)
                if match:
                    date_str = match.group(1).strip()
                    result['invoiceDate'] = convert_date_format(date_str)
                    break
            
            # Extract Terms of Payment
            pattern = r'Terms?\s*(?:of\s*)?Payment[.\s:]*([^\n]+?)(?:\n|$)'
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                result['termsOfPayment'] = match.group(1).strip()
            
            # Extract Terms of Delivery
            pattern = r'Terms?\s*(?:of\s*)?Delivery[.\s:]*([^\n]+?)(?:\n|$)'
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                result['termsOfDelivery'] = match.group(1).strip()
            
            # Extract Currency
            currency_pattern = r'(USD|EUR|INR|GBP|JPY|CHF|AUD|CAD|SGD|HKD)'
            match = re.search(currency_pattern, full_text)
            if match:
                result['currency'] = match.group(1)
            
            # Extract Amount
            patterns = [
                r'Grand\s*Total\s+(\d+\.?\d*)',
                r'Total\s*Amount[.\s:]*(\d+\.?\d*)',
                r'Total[.\s:]*(\d+\.?\d*)'
            ]
            for pattern in patterns:
                match = re.search(pattern, full_text, re.IGNORECASE)
                if match:
                    result['amount'] = match.group(1)
                    break
            
            return result
    
    except Exception as e:
        return {'error': f'PDF parsing error: {str(e)}'}

def extract_shipment_data(pdf_path):
    """Extract shipping bill data from PDF including Invoice Number"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"
            
            # Print first 1000 chars for debugging
            print("🔍 ===== SHIPPING BILL TEXT (first 1000 chars) =====")
            print(full_text[:1000])
            print("🔍 =================================================")
            
            result = {
                'shipBillNo': '',
                'shipBillingDate': '',
                'invoiceNumber': ''
            }
            
            # ===== Extract Ship Bill No (7-digit number) =====
            sb_patterns = [
                r'SB\s*No[.:\s]*(\d{7})',
                r'SB\s*NO[.:\s]*(\d{7})',
                r'Shipping\s*Bill\s*No[.:\s]*(\d{7})',
                r'SB\s*Number[.:\s]*(\d{7})',
                r'SB\s*NO\s*(\d{7})',
                r'Port Code.*?(\d{7})',
                r'\b(\d{7})\b',
            ]
            for pattern in sb_patterns:
                match = re.search(pattern, full_text[:500], re.IGNORECASE)
                if match:
                    potential_sb = match.group(1).strip()
                    # Make sure it's a valid 7-digit number
                    if len(potential_sb) == 7 and potential_sb.isdigit():
                        result['shipBillNo'] = potential_sb
                        print(f"🔍 Found Ship Bill No: {result['shipBillNo']}")
                        break
            
            # ===== Extract Ship Bill Date =====
            date_patterns = [
                r'SB\s*Date[.:\s]*(\d{1,2})\s*[-/]?\s*([A-Z]{3})\s*[-/]?\s*(\d{2,4})',
                r'SB\s*DATE[.:\s]*(\d{1,2})\s*[-/]?\s*([A-Z]{3})\s*[-/]?\s*(\d{2,4})',
                r'Shipping\s*Bill\s*Date[.:\s]*(\d{1,2})\s*[-/]?\s*([A-Z]{3})\s*[-/]?\s*(\d{2,4})',
                r'(\d{1,2})\s*[-/]?\s*([A-Z]{3})\s*[-/]?\s*(\d{2,4})',
            ]
            for pattern in date_patterns:
                match = re.search(pattern, full_text, re.IGNORECASE)
                if match:
                    day, month_str, year = match.groups()
                    month_map = {
                        'JAN': '01', 'FEB': '02', 'MAR': '03', 'APR': '04',
                        'MAY': '05', 'JUN': '06', 'JUL': '07', 'AUG': '08',
                        'SEP': '09', 'OCT': '10', 'NOV': '11', 'DEC': '12'
                    }
                    month = month_map.get(month_str.upper(), '')
                    if month:
                        year_int = int(year)
                        if year_int < 100:
                            year_int = 2000 + year_int if year_int < 50 else 1900 + year_int
                        result['shipBillingDate'] = f"{int(day):02d}-{month}-{year_int}"
                        print(f"🔍 Found Ship Bill Date: {result['shipBillingDate']}")
                        break
            
            # ===== Extract Invoice Number (10-digit number starting with 25) =====
            print("🔍 Searching for Invoice Number...")
            
            # Look for INV NO pattern
            inv_patterns = [
                r'INV\s*NO[.:\s]*(\d{10})',
                r'INV\s*NO[.:\s]*(\d{9,12})',
                r'INVOICE\s*NO[.:\s]*(\d{10})',
                r'INV\.?\s*NO[.:\s]*(\d{10})',
                r'INV\s*NO\s*[:]?\s*(\d{10})',
                r'Invoice\s*Number[.:\s]*(\d{10})',
                r'INV\s*NO\s*(\d{10})',
                r'INV\s*NO\s*\|\s*(\d{10})',
                r'INV\s*NO\s*(\d{10})',
                r'INV\s*NO:\s*(\d{10})',
                r'INV\s*NO\s+(\d{10})',
                r'Invoice\s+No\s+(\d{10})',
                r'INV\s*NO\s*=\s*(\d{10})',
            ]
            
            for pattern in inv_patterns:
                match = re.search(pattern, full_text, re.IGNORECASE)
                if match:
                    potential_inv = match.group(1).strip()
                    if len(potential_inv) >= 9 and potential_inv.isdigit():
                        result['invoiceNumber'] = potential_inv
                        print(f"🔍 Found Invoice Number from pattern: {result['invoiceNumber']}")
                        break
            
            # If still not found, look for 10-digit number starting with 25
            if not result['invoiceNumber']:
                # Find all 10-digit numbers
                all_numbers = re.findall(r'\b(\d{10})\b', full_text)
                for num in all_numbers:
                    if num.startswith('25'):
                        result['invoiceNumber'] = num
                        print(f"🔍 Found Invoice Number (10-digit starting with 25): {result['invoiceNumber']}")
                        break
            
            # If still not found, look for any 10-digit number
            if not result['invoiceNumber']:
                # Look for INV or Invoice near a number
                inv_section = re.search(r'INV[OI]CE?\s*.*?(\d{10})', full_text, re.IGNORECASE)
                if inv_section:
                    result['invoiceNumber'] = inv_section.group(1)
                    print(f"🔍 Found Invoice Number from INV section: {result['invoiceNumber']}")
            
            print(f"🔍 Final Result: {result}")
            return result
    
    except Exception as e:
        print(f"✗ Error extracting shipment data: {e}")
        import traceback
        traceback.print_exc()
        return {'error': f'PDF parsing error: {str(e)}'}

# ============================================================================
# AUTHENTICATION ROUTES WITH EMAIL VERIFICATION
# ============================================================================

@app.route('/')
def index():
    """Serve the main HTML page if logged in, otherwise login page"""
    if 'user_id' in session:
        user = get_current_user()
        if user and user.get('email_verified', False):
            return send_file('index.html', mimetype='text/html')
        else:
            session.clear()
            return redirect('/login')
    return send_file('login.html', mimetype='text/html')

@app.route('/login')
def login_page():
    if 'user_id' in session:
        return redirect('/')
    return send_file('login.html', mimetype='text/html')

@app.route('/register')
def register_page():
    if 'user_id' in session:
        return redirect('/')
    return send_file('register.html', mimetype='text/html')

@app.route('/verify-otp')
def verify_otp_page():
    """Serve OTP verification page"""
    email = request.args.get('email')
    if not email:
        return redirect('/register')
    return send_file('verify_otp.html', mimetype='text/html')

@app.route('/forgot-password')
def forgot_password_page():
    """Serve forgot password page"""
    return send_file('forgot_password.html', mimetype='text/html')

@app.route('/reset-password')
def reset_password_page():
    """Serve reset password page"""
    email = request.args.get('email')
    if not email:
        return redirect('/forgot-password')
    return send_file('reset_password.html', mimetype='text/html')

# ============================================================================
# AUTHENTICATION API ROUTES
# ============================================================================

@app.route('/api/send-otp', methods=['POST'])
def send_otp():
    """Send OTP to email for verification"""
    try:
        data = request.get_json()
        email = data.get('email', '').strip()
        username = data.get('username', '').strip()
        
        if not email:
            return jsonify({'status': 'error', 'message': 'Email is required'}), 400
        
        if not username or len(username) < 3:
            return jsonify({'status': 'error', 'message': 'Username must be at least 3 characters'}), 400
        
        # Check if email is already registered and verified
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT id, email_verified FROM users WHERE email = %s', (email,))
        existing = cursor.fetchone()
        
        if existing:
            if existing[1]:  # email_verified is True
                cursor.close()
                return_db_connection(conn)
                return jsonify({'status': 'error', 'message': 'Email already registered and verified'}), 400
            else:
                # User exists but not verified - we'll allow resending OTP
                pass
        
        cursor.close()
        return_db_connection(conn)
        
        # Generate OTP
        otp = generate_otp()
        
        # Store OTP in database
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Delete any existing unverified OTPs for this email
            cursor.execute('DELETE FROM email_verifications WHERE email = %s AND is_used = FALSE', (email,))
            
            # Insert new OTP
            expires_at = datetime.now() + timedelta(minutes=OTP_EXPIRY_MINUTES)
            cursor.execute('''
                INSERT INTO email_verifications (email, otp, expires_at)
                VALUES (%s, %s, %s)
            ''', (email, otp, expires_at))
            
            conn.commit()
            
            # Send OTP via email (non-blocking)
            send_otp_email(email, otp, username)
            
            return jsonify({
                'status': 'success',
                'message': f'OTP sent to {email}',
                'email': email
            })
            
        except Exception as e:
            conn.rollback()
            return jsonify({'status': 'error', 'message': str(e)}), 500
        finally:
            cursor.close()
            return_db_connection(conn)
            
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/verify-otp', methods=['POST'])
def verify_otp():
    """Verify OTP and complete registration"""
    try:
        data = request.get_json()
        email = data.get('email', '').strip()
        otp = data.get('otp', '').strip()
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        if not email or not otp:
            return jsonify({'status': 'error', 'message': 'Email and OTP are required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        try:
            # Verify OTP
            cursor.execute('''
                SELECT id, otp, created_at, expires_at, is_used
                FROM email_verifications
                WHERE email = %s AND otp = %s AND is_used = FALSE
                ORDER BY created_at DESC
                LIMIT 1
            ''', (email, otp))
            
            verification = cursor.fetchone()
            
            if not verification:
                return jsonify({'status': 'error', 'message': 'Invalid OTP'}), 400
            
            # Check if OTP expired
            if verification['expires_at'] < datetime.now():
                return jsonify({'status': 'error', 'message': 'OTP has expired. Please request a new one.'}), 400
            
            # Mark OTP as used
            cursor.execute('UPDATE email_verifications SET is_used = TRUE WHERE id = %s', (verification['id'],))
            
            # Check if user already exists
            cursor.execute('SELECT id FROM users WHERE email = %s', (email,))
            existing_user = cursor.fetchone()
            
            if existing_user:
                # Update existing user's verification status
                cursor.execute('''
                    UPDATE users 
                    SET email_verified = TRUE, username = %s, password_hash = %s
                    WHERE email = %s
                    RETURNING id
                ''', (username, generate_password_hash(password), email))
                
                user_id = cursor.fetchone()['id']
            else:
                # Create new user
                password_hash = generate_password_hash(password)
                cursor.execute('''
                    INSERT INTO users (username, email, password_hash, email_verified)
                    VALUES (%s, %s, %s, TRUE)
                    RETURNING id
                ''', (username, email, password_hash))
                
                user_id = cursor.fetchone()['id']
            
            conn.commit()
            
            # Auto-login after verification
            session.permanent = True
            session['user_id'] = user_id
            session['username'] = username
            
            return jsonify({
                'status': 'success',
                'message': 'Email verified successfully!',
                'user': {
                    'id': user_id,
                    'username': username,
                    'email': email
                }
            })
            
        except Exception as e:
            conn.rollback()
            return jsonify({'status': 'error', 'message': str(e)}), 500
        finally:
            cursor.close()
            return_db_connection(conn)
            
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/login', methods=['POST'])
def login():
    """Login user - only if email is verified"""
    try:
        data = request.get_json()
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        if not username or not password:
            return jsonify({'status': 'error', 'message': 'Username and password required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        try:
            cursor.execute('''
                SELECT id, username, email, password_hash, email_verified, is_active 
                FROM users 
                WHERE username = %s OR email = %s
            ''', (username, username))
            
            user = cursor.fetchone()
            
            if not user:
                return jsonify({'status': 'error', 'message': 'Invalid credentials'}), 401
            
            if not user['is_active']:
                return jsonify({'status': 'error', 'message': 'Account is disabled'}), 401
            
            # Check if email is verified
            if not user['email_verified']:
                return jsonify({
                    'status': 'error', 
                    'message': 'Please verify your email first. Check your inbox for OTP.',
                    'needs_verification': True,
                    'email': user['email']
                }), 401
            
            if not check_password_hash(user['password_hash'], password):
                return jsonify({'status': 'error', 'message': 'Invalid credentials'}), 401
            
            # Update last login
            cursor.execute('UPDATE users SET last_login = CURRENT_TIMESTAMP WHERE id = %s', (user['id'],))
            conn.commit()
            
            session.permanent = True
            session['user_id'] = user['id']
            session['username'] = user['username']
            
            return jsonify({
                'status': 'success',
                'message': 'Login successful!',
                'user': {
                    'id': user['id'],
                    'username': user['username'],
                    'email': user['email']
                }
            })
            
        except Exception as e:
            conn.rollback()
            return jsonify({'status': 'error', 'message': str(e)}), 500
        finally:
            cursor.close()
            return_db_connection(conn)
            
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/resend-otp', methods=['POST'])
def resend_otp():
    """Resend OTP to email"""
    try:
        data = request.get_json()
        email = data.get('email', '').strip()
        
        if not email:
            return jsonify({'status': 'error', 'message': 'Email is required'}), 400
        
        # Generate new OTP
        otp = generate_otp()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Delete existing unverified OTPs
            cursor.execute('DELETE FROM email_verifications WHERE email = %s AND is_used = FALSE', (email,))
            
            # Insert new OTP
            expires_at = datetime.now() + timedelta(minutes=OTP_EXPIRY_MINUTES)
            cursor.execute('''
                INSERT INTO email_verifications (email, otp, expires_at)
                VALUES (%s, %s, %s)
            ''', (email, otp, expires_at))
            
            conn.commit()
            
            # Send OTP (non-blocking)
            send_otp_email(email, otp, 'User')
            
            return jsonify({
                'status': 'success',
                'message': 'OTP resent successfully'
            })
            
        except Exception as e:
            conn.rollback()
            return jsonify({'status': 'error', 'message': str(e)}), 500
        finally:
            cursor.close()
            return_db_connection(conn)
            
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'status': 'success', 'message': 'Logged out successfully'})

@app.route('/api/check-auth', methods=['GET'])
def check_auth():
    if 'user_id' in session:
        user = get_current_user()
        if user and user.get('email_verified', False):
            return jsonify({
                'authenticated': True,
                'user': {
                    'id': user['id'],
                    'username': user['username'],
                    'email': user['email']
                }
            })
        else:
            session.clear()
    return jsonify({'authenticated': False}), 401

@app.route('/api/current-user', methods=['GET'])
@login_required
def current_user():
    """Get current user info"""
    user = get_current_user()
    if user:
        return jsonify({
            'status': 'success',
            'user': {
                'id': user['id'],
                'username': user['username'],
                'email': user['email']
            }
        })
    return jsonify({'status': 'error', 'message': 'User not found'}), 404

# ============================================================================
# INVOICE ENDPOINTS
# ============================================================================

@app.route('/extract-invoice', methods=['POST'])
@login_required
def extract_invoice():
    """Extract invoice data from uploaded PDF"""
    if 'file' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file uploaded'}), 400
    
    file = request.files['file']
    
    if file.filename == '' or not allowed_file(file.filename):
        return jsonify({'status': 'error', 'message': 'Invalid file. Only PDF allowed'}), 400
    
    try:
        filename = secure_filename(file.filename)
        filepath = os.path.join(UPLOAD_FOLDER, f"{get_user_id()}_{filename}")
        file.save(filepath)
        
        extracted_data = extract_invoice_data(filepath)
        os.remove(filepath)
        
        if 'error' in extracted_data:
            return jsonify({'status': 'error', 'message': extracted_data['error']}), 400
        
        return jsonify({
            'status': 'success',
            'data': extracted_data
        })
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/extract-invoice-bundle', methods=['POST'])
@login_required
def extract_invoice_bundle():
    """Extract invoice data from MULTIPLE uploaded PDFs"""
    if 'files' not in request.files:
        return jsonify({'status': 'error', 'message': 'No files uploaded'}), 400
    
    files = request.files.getlist('files')
    
    if not files or len(files) == 0:
        return jsonify({'status': 'error', 'message': 'No files selected'}), 400
    
    try:
        all_extracted_data = []
        
        for file in files:
            if file.filename == '' or not allowed_file(file.filename):
                continue
            
            filename = secure_filename(file.filename)
            filepath = os.path.join(UPLOAD_FOLDER, f"{get_user_id()}_{filename}")
            file.save(filepath)
            
            extracted_data = extract_invoice_data(filepath)
            os.remove(filepath)
            
            if 'error' not in extracted_data:
                extracted_data['fileName'] = file.filename
                all_extracted_data.append(extracted_data)
        
        if len(all_extracted_data) == 0:
            return jsonify({'status': 'error', 'message': 'No valid PDF files found'}), 400
        
        return jsonify({
            'status': 'success',
            'data': all_extracted_data,
            'count': len(all_extracted_data)
        })
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/submit-invoice', methods=['POST'])
@login_required
def submit_invoice():
    """Store submitted invoice data in database"""
    try:
        data = request.get_json()
        user_id = get_user_id()
        
        if not data.get('invoiceNumber') or not data.get('invoiceDate'):
            return jsonify({'status': 'error', 'message': 'Invoice Number and Date are required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                INSERT INTO invoices 
                (user_id, invoice_number, invoice_date, terms_of_payment, terms_of_delivery, currency, amount)
                VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id
            ''', (
                user_id,
                data.get('invoiceNumber', ''),
                data.get('invoiceDate', ''),
                data.get('termsOfPayment', ''),
                data.get('termsOfDelivery', ''),
                data.get('currency', ''),
                data.get('amount', '')
            ))
            
            invoice_id = cursor.fetchone()[0]
            conn.commit()
            
            return jsonify({
                'status': 'success',
                'message': 'Invoice submitted successfully',
                'id': invoice_id
            })
            
        except Exception as e:
            conn.rollback()
            return jsonify({'status': 'error', 'message': str(e)}), 500
        finally:
            cursor.close()
            return_db_connection(conn)
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/get-invoices', methods=['GET'])
@login_required
def get_invoices():
    """Get all submitted invoices for current user"""
    try:
        user_id = get_user_id()
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute('''
            SELECT id, invoice_number, invoice_date, terms_of_payment, 
                   terms_of_delivery, currency, amount, submitted_at 
            FROM invoices 
            WHERE user_id = %s 
            ORDER BY submitted_at DESC
        ''', (user_id,))
        
        rows = cursor.fetchall()
        cursor.close()
        return_db_connection(conn)
        
        invoices = []
        for row in rows:
            invoices.append({
                'id': row['id'],
                'invoiceNumber': row['invoice_number'],
                'invoiceDate': row['invoice_date'],
                'termsOfPayment': row['terms_of_payment'],
                'termsOfDelivery': row['terms_of_delivery'],
                'currency': row['currency'],
                'amount': row['amount'],
                'submittedAt': row['submitted_at']
            })
        
        return jsonify({'status': 'success', 'data': invoices})
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ============================================================================
# SHIPMENT ENDPOINTS
# ============================================================================

@app.route('/extract-shipment', methods=['POST'])
@login_required
def extract_shipment():
    """Extract shipment data from uploaded PDF including Invoice Number"""
    if 'file' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file uploaded'}), 400
    
    file = request.files['file']
    
    if file.filename == '' or not allowed_file(file.filename):
        return jsonify({'status': 'error', 'message': 'Invalid file. Only PDF allowed'}), 400
    
    try:
        filename = secure_filename(file.filename)
        filepath = os.path.join(UPLOAD_FOLDER, f"{get_user_id()}_{filename}")
        file.save(filepath)
        
        extracted_data = extract_shipment_data(filepath)
        os.remove(filepath)
        
        if 'error' in extracted_data:
            return jsonify({'status': 'error', 'message': extracted_data['error']}), 400
        
        return jsonify({
            'status': 'success',
            'data': extracted_data
        })
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/extract-shipment-bundle', methods=['POST'])
@login_required
def extract_shipment_bundle():
    """Extract shipment data from MULTIPLE uploaded PDFs including Invoice Numbers"""
    if 'files' not in request.files:
        return jsonify({'status': 'error', 'message': 'No files uploaded'}), 400
    
    files = request.files.getlist('files')
    
    if not files or len(files) == 0:
        return jsonify({'status': 'error', 'message': 'No files selected'}), 400
    
    try:
        all_extracted_data = []
        
        for file in files:
            if file.filename == '' or not allowed_file(file.filename):
                continue
            
            filename = secure_filename(file.filename)
            filepath = os.path.join(UPLOAD_FOLDER, f"{get_user_id()}_{filename}")
            file.save(filepath)
            
            extracted_data = extract_shipment_data(filepath)
            os.remove(filepath)
            
            if 'error' not in extracted_data:
                extracted_data['fileName'] = file.filename
                all_extracted_data.append(extracted_data)
        
        if len(all_extracted_data) == 0:
            return jsonify({'status': 'error', 'message': 'No valid PDF files found'}), 400
        
        return jsonify({
            'status': 'success',
            'data': all_extracted_data,
            'count': len(all_extracted_data)
        })
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/submit-shipment', methods=['POST'])
@login_required
def submit_shipment():
    """Store submitted shipment data in database with invoice_number"""
    try:
        data = request.get_json()
        user_id = get_user_id()
        
        if not data.get('shipBillNo') or not data.get('shipBillingDate'):
            return jsonify({'status': 'error', 'message': 'Ship Bill No and Date are required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                INSERT INTO shipments 
                (user_id, ship_bill_no, ship_billing_date, invoice_number)
                VALUES (%s, %s, %s, %s) RETURNING id
            ''', (
                user_id,
                data.get('shipBillNo', ''),
                data.get('shipBillingDate', ''),
                data.get('invoiceNumber', '')  # NEW: Store the invoice number
            ))
            
            shipment_id = cursor.fetchone()[0]
            conn.commit()
            
            return jsonify({
                'status': 'success',
                'message': 'Shipment submitted successfully',
                'id': shipment_id
            })
            
        except Exception as e:
            conn.rollback()
            return jsonify({'status': 'error', 'message': str(e)}), 500
        finally:
            cursor.close()
            return_db_connection(conn)
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/get-shipments', methods=['GET'])
@login_required
def get_shipments():
    """Get all submitted shipments for current user"""
    try:
        user_id = get_user_id()
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute('''
            SELECT id, ship_bill_no, ship_billing_date, invoice_number, submitted_at 
            FROM shipments 
            WHERE user_id = %s 
            ORDER BY submitted_at DESC
        ''', (user_id,))
        
        rows = cursor.fetchall()
        cursor.close()
        return_db_connection(conn)
        
        shipments = []
        for row in rows:
            shipments.append({
                'id': row['id'],
                'shipBillNo': row['ship_bill_no'],
                'shipBillingDate': row['ship_billing_date'],
                'invoiceNumber': row.get('invoice_number', ''),
                'submittedAt': row['submitted_at']
            })
        
        return jsonify({'status': 'success', 'data': shipments})
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ============================================================================
# EXPORT ENDPOINT - MATCHES INVOICE WITH SHIPMENT BY INVOICE NUMBER
# ============================================================================

@app.route('/export-combined', methods=['GET'])
@login_required
def export_combined():
    """Export all invoices with matched shipments by invoice number"""
    try:
        user_id = get_user_id()
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Get all invoices for current user
        cursor.execute('''
            SELECT id, invoice_number, invoice_date, terms_of_payment, 
                   terms_of_delivery, currency, amount 
            FROM invoices 
            WHERE user_id = %s 
            ORDER BY id
        ''', (user_id,))
        invoice_rows = cursor.fetchall()
        
        # Get all shipments for current user
        cursor.execute('''
            SELECT id, ship_bill_no, ship_billing_date, invoice_number 
            FROM shipments 
            WHERE user_id = %s 
            ORDER BY id
        ''', (user_id,))
        shipment_rows = cursor.fetchall()
        
        cursor.close()
        return_db_connection(conn)
        
        # Create mapping: invoice_number -> shipment data
        shipment_map = {}
        for ship in shipment_rows:
            inv_num = ship.get('invoice_number', '')
            if inv_num:
                if inv_num not in shipment_map:
                    shipment_map[inv_num] = {
                        'shipBillNo': ship['ship_bill_no'],
                        'shipBillingDate': ship['ship_billing_date']
                    }
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "All Data"
        
        current_row = 1
        
        # ===== HEADER ROW - NO SHIP S.NO =====
        headers = ['S.No', 'Invoice Number', 'Invoice Date', 'Terms of Payment', 
                   'Terms of Delivery', 'Currency', 'Amount',
                   'Ship Bill No', 'Ship Billing Date']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="1b5e20", end_color="1b5e20", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        current_row += 1
        
        # ===== DATA ROWS =====
        sno = 1
        for invoice in invoice_rows:
            inv_num = invoice['invoice_number']
            
            # Check if this invoice has a matching shipment
            ship_data = shipment_map.get(inv_num, {})
            
            row_data = [
                sno,
                inv_num,
                invoice['invoice_date'],
                invoice['terms_of_payment'] or '',
                invoice['terms_of_delivery'] or '',
                invoice['currency'] or '',
                invoice['amount'] or '',
                ship_data.get('shipBillNo', ''),
                ship_data.get('shipBillingDate', '')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            sno += 1
            current_row += 1
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        user = get_current_user()
        username = user['username'] if user else 'user'
        
        return send_file(
            file_stream,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{username}_combined_data.xlsx'
        )
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
        
# ============================================================================
# ERROR HANDLERS
# ============================================================================

@app.errorhandler(404)
def not_found(error):
    return jsonify({'status': 'error', 'message': 'Endpoint not found'}), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({'status': 'error', 'message': 'Internal server error'}), 500

# ============================================================================
# USERNAME & EMAIL VALIDATION ROUTES
# ============================================================================

@app.route('/api/check-username', methods=['POST'])
def check_username():
    """Check if username is already taken"""
    try:
        data = request.get_json()
        username = data.get('username', '').strip()
        
        if not username:
            return jsonify({'status': 'error', 'message': 'Username is required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT id FROM users WHERE username = %s', (username,))
        existing = cursor.fetchone()
        
        cursor.close()
        return_db_connection(conn)
        
        if existing:
            return jsonify({'status': 'error', 'message': 'Username already taken'}), 409
        else:
            return jsonify({'status': 'success', 'message': 'Username available'})
            
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/check-email', methods=['POST'])
def check_email():
    """Check if email is already registered"""
    try:
        data = request.get_json()
        email = data.get('email', '').strip()
        
        if not email:
            return jsonify({'status': 'error', 'message': 'Email is required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT id, email_verified FROM users WHERE email = %s', (email,))
        existing = cursor.fetchone()
        
        cursor.close()
        return_db_connection(conn)
        
        if existing:
            if existing[1]:  # email_verified is True
                return jsonify({'status': 'error', 'message': 'Email already registered'}), 409
            else:
                return jsonify({'status': 'error', 'message': 'Email already registered but not verified. Please verify your email.'}), 409
        else:
            return jsonify({'status': 'success', 'message': 'Email available'})
            
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ============================================================================
# FORGOT PASSWORD ROUTES
# ============================================================================

@app.route('/api/request-password-reset', methods=['POST'])
def request_password_reset():
    """Send OTP for password reset"""
    try:
        data = request.get_json()
        email = data.get('email', '').strip()
        
        if not email:
            return jsonify({'status': 'error', 'message': 'Email is required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if user exists with this email
        cursor.execute('SELECT id, username FROM users WHERE email = %s AND email_verified = TRUE', (email,))
        user = cursor.fetchone()
        
        if not user:
            cursor.close()
            return_db_connection(conn)
            return jsonify({'status': 'error', 'message': 'No account found with this email'}), 404
        
        user_id = user[0]
        username = user[1]
        
        # Generate OTP
        otp = generate_otp()
        
        # Delete any existing unverified OTPs for this email
        cursor.execute('DELETE FROM email_verifications WHERE email = %s AND is_used = FALSE', (email,))
        
        # Insert new OTP
        expires_at = datetime.now() + timedelta(minutes=OTP_EXPIRY_MINUTES)
        cursor.execute('''
            INSERT INTO email_verifications (email, otp, expires_at)
            VALUES (%s, %s, %s)
        ''', (email, otp, expires_at))
        
        conn.commit()
        cursor.close()
        return_db_connection(conn)
        
        # Send OTP via email with reset instructions (non-blocking)
        send_password_reset_email(email, otp, username)
        
        return jsonify({
            'status': 'success',
            'message': 'Password reset OTP sent to your email',
            'email': email
        })
        
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/verify-reset-otp', methods=['POST'])
def verify_reset_otp():
    """Verify OTP and reset password"""
    try:
        data = request.get_json()
        email = data.get('email', '').strip()
        otp = data.get('otp', '').strip()
        new_password = data.get('newPassword', '')
        
        if not email or not otp or not new_password:
            return jsonify({'status': 'error', 'message': 'Email, OTP, and new password are required'}), 400
        
        if len(new_password) < 6:
            return jsonify({'status': 'error', 'message': 'Password must be at least 6 characters'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verify OTP
        cursor.execute('''
            SELECT id, created_at, expires_at, is_used
            FROM email_verifications
            WHERE email = %s AND otp = %s AND is_used = FALSE
            ORDER BY created_at DESC
            LIMIT 1
        ''', (email, otp))
        
        verification = cursor.fetchone()
        
        if not verification:
            cursor.close()
            return_db_connection(conn)
            return jsonify({'status': 'error', 'message': 'Invalid OTP'}), 400
        
        # Check if OTP expired
        if verification[2] < datetime.now():
            cursor.close()
            return_db_connection(conn)
            return jsonify({'status': 'error', 'message': 'OTP has expired. Please request a new one.'}), 400
        
        # Mark OTP as used
        cursor.execute('UPDATE email_verifications SET is_used = TRUE WHERE id = %s', (verification[0],))
        
        # Update password
        password_hash = generate_password_hash(new_password)
        cursor.execute('UPDATE users SET password_hash = %s WHERE email = %s', (password_hash, email))
        
        conn.commit()
        cursor.close()
        return_db_connection(conn)
        
        return jsonify({
            'status': 'success',
            'message': 'Password reset successfully! Please login with your new password.'
        })
        
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ============================================================================
# DATABASE INITIALIZATION ON STARTUP (FOR PRODUCTION)
# ============================================================================

# This runs when Gunicorn starts the app on Railway
with app.app_context():
    try:
        init_db()
        print("✓ Database initialized successfully on startup")
    except Exception as e:
        print(f"✗ Database initialization failed: {e}")

# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

if __name__ == '__main__':
    # This runs when you run python app.py locally
    app.run(debug=True, host='0.0.0.0', port=5000)