import os
import re
import sqlite3
import json
from datetime import datetime
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

app = Flask(__name__)
CORS(app)

# Configuration
UPLOAD_FOLDER = 'uploads'
DATABASE = 'parser.db'
ALLOWED_EXTENSIONS = {'pdf'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ============================================================================
# DATABASE INITIALIZATION
# ============================================================================

def init_db():
    """Initialize SQLite database with required tables"""
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    
    # Invoices table
    c.execute('''
        CREATE TABLE IF NOT EXISTS invoices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_number TEXT NOT NULL,
            invoice_date TEXT NOT NULL,
            terms_of_payment TEXT,
            terms_of_delivery TEXT,
            currency TEXT,
            amount TEXT,
            submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Shipments table
    c.execute('''
        CREATE TABLE IF NOT EXISTS shipments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ship_bill_no TEXT NOT NULL,
            ship_billing_date TEXT NOT NULL,
            submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    conn.commit()
    conn.close()

init_db()

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def allowed_file(filename):
    """Check if file has allowed extension"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def convert_date_format(date_str):
    """
    Convert various date formats to DD-MM-YYYY format
    Handles: DD.MM.YYYY, DD/MM/YYYY, DD-MM-YYYY, DD-MMM-YY (for shipping bills)
    """
    if not date_str:
        return ""
    
    date_str = date_str.strip()
    
    # Handle DD-MMM-YY format (e.g., 03-SEP-25)
    month_map = {
        'JAN': '01', 'FEB': '02', 'MAR': '03', 'APR': '04', 'MAY': '05', 'JUN': '06',
        'JUL': '07', 'AUG': '08', 'SEP': '09', 'OCT': '10', 'NOV': '11', 'DEC': '12'
    }
    
    # Try DD-MMM-YY format
    match = re.search(r'(\d{1,2})\s*[-/]?\s*([A-Z]{3})\s*[-/]?\s*(\d{2,4})', date_str, re.IGNORECASE)
    if match:
        day, month_str, year = match.groups()
        month_str_upper = month_str.upper()
        if month_str_upper in month_map:
            month = month_map[month_str_upper]
            # Convert 2-digit year to 4-digit
            year = int(year)
            if year < 100:
                year = 2000 + year if year < 50 else 1900 + year
            return f"{int(day):02d}-{month}-{year}"
    
    # Try DD.MM.YYYY or DD/MM/YYYY format
    match = re.search(r'(\d{1,2})\s*[./]\s*(\d{1,2})\s*[./]\s*(\d{4})', date_str)
    if match:
        day, month, year = match.groups()
        return f"{int(day):02d}-{int(month):02d}-{year}"
    
    # Try DD-MM-YYYY format
    match = re.search(r'(\d{1,2})\s*-\s*(\d{1,2})\s*-\s*(\d{4})', date_str)
    if match:
        day, month, year = match.groups()
        return f"{int(day):02d}-{int(month):02d}-{year}"
    
    return date_str

def extract_invoice_data(pdf_path):
    """
    Extract invoice data from PDF using pdfplumber
    Returns dict with extracted fields
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Extract text from all pages
            full_text = ""
            for page in pdf.pages:
                full_text += page.extract_text() + "\n"

            # DEBUG: Print what we got
            print("PDF TEXT EXTRACTED:")
            print(full_text[:500])  # Print first 500 characters
            print("=" * 50)

            # Initialize result dictionary
            result = {
                'invoiceNumber': '',
                'invoiceDate': '',
                'termsOfPayment': '',
                'termsOfDelivery': '',
                'currency': '',
                'amount': ''
            }
            
            # Extract Invoice Number - NOW HANDLES LINE BREAKS AND COMPANY NAME
            patterns = [
                r'Invoice\s*No\s*&\s*Date\s*\n.*?(\d{10})',  # Handles newline + company name + number
                r'Invoice\s*No\s*&\s*Date\s*[^\d]*(\d+)',     # Any non-digits between label and number
                r'Invoice\s*No[.\s:]+(\d+)',
                r'INV[OICE]*\s*No[.\s:]+(\d+)',
            ]

            for pattern in patterns:
                match = re.search(pattern, full_text, re.IGNORECASE | re.DOTALL)
                if match:
                    result['invoiceNumber'] = match.group(1).strip()
                    break
            
            # Extract Invoice Date
            patterns = [
                r'Invoice\s*(?:No\s*&\s*)?Date[.\s:]*([^\n]+?)(?:\n|$)',
                r'Dt:\s*(\d{1,2}[./\-]\d{1,2}[./\-]\d{2,4})',
                r'Date[.\s:]*(\d{1,2}[./\-]\d{1,2}[./\-]\d{2,4})'
            ]
            for pattern in patterns:
                match = re.search(pattern, full_text, re.IGNORECASE)
                if match:
                    date_str = match.group(1).strip()
                    result['invoiceDate'] = convert_date_format(date_str)
                    break
            
            # Extract Terms of Payment
            pattern = r'Terms?\s*(?:of\s*)?Payment[.\s:]*([^\n]+?)(?:\n|$)'
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                result['termsOfPayment'] = match.group(1).strip()
            
            # Extract Terms of Delivery
            pattern = r'Terms?\s*(?:of\s*)?Delivery[.\s:]*([^\n]+?)(?:\n|$)'
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                result['termsOfDelivery'] = match.group(1).strip()
            
            # Extract Currency
            currency_pattern = r'(USD|EUR|INR|GBP|JPY|CHF|AUD|CAD|SGD|HKD)'
            match = re.search(currency_pattern, full_text)
            if match:
                result['currency'] = match.group(1)
            
            # Extract Amount
            patterns = [
                r'Grand\s*Total\s+(\d+\.?\d*)',
                r'Total\s*Amount[.\s:]*(\d+\.?\d*)',
                r'Total[.\s:]*(\d+\.?\d*)'
            ]
            for pattern in patterns:
                match = re.search(pattern, full_text, re.IGNORECASE)
                if match:
                    result['amount'] = match.group(1)
                    break
            
            return result
    
    except Exception as e:
        return {'error': f'PDF parsing error: {str(e)}'}

def extract_shipment_data(pdf_path):
    """
    Extract shipping bill data from PDF - DEBUG VERSION
    """
    try:
        print("\n" + "="*60)
        print("STARTING SHIPMENT EXTRACTION")
        print("="*60)
        
        with pdfplumber.open(pdf_path) as pdf:
            print(f"PDF opened successfully. Pages: {len(pdf.pages)}")
            
            first_page_text = pdf.pages[0].extract_text()
            
            print("\nFIRST 1500 CHARACTERS OF PDF:")
            print(first_page_text[:1500])
            print("="*60)
            
            result = {
                'shipBillNo': '',
                'shipBillingDate': ''
            }
            
            # Get only the first 2000 characters (top section with SB info)
            top_section = first_page_text[:2000]
            
            # Extract any 7-digit number (SB No is usually 7 digits)
            print("\nSearching for 7-digit Ship Bill No...")
            match = re.search(r'\b(\d{7})\b', top_section)
            if match:
                result['shipBillNo'] = match.group(1).strip()
                print(f"✓ FOUND SHIP BILL NO: {result['shipBillNo']}")
            else:
                print("✗ NO SHIP BILL NO FOUND - searching for any digits...")
                # Try broader pattern
                match = re.search(r'(\d+)', top_section)
                if match:
                    print(f"  Found numbers in text: {match.group(1)}")
            
            # Extract date in DD-MMM-YY format
            print("\nSearching for date (DD-MMM-YY format)...")
            match = re.search(r'(\d{1,2})-([A-Z]{3})-(\d{2,4})', top_section, re.IGNORECASE)
            if match:
                day, month_str, year = match.groups()
                print(f"  Found date components: {day}-{month_str}-{year}")
                month_map = {
                    'JAN': '01', 'FEB': '02', 'MAR': '03', 'APR': '04',
                    'MAY': '05', 'JUN': '06', 'JUL': '07', 'AUG': '08',
                    'SEP': '09', 'OCT': '10', 'NOV': '11', 'DEC': '12'
                }
                month = month_map.get(month_str.upper(), '')
                year_int = int(year)
                if year_int < 100:
                    year_int = 2000 + year_int if year_int < 50 else 1900 + year_int
                result['shipBillingDate'] = f"{int(day):02d}-{month}-{year_int}"
                print(f"✓ FOUND SHIP BILL DATE: {result['shipBillingDate']}")
            else:
                print("✗ NO SHIP BILL DATE FOUND")
            
            print("\nFINAL RESULT:")
            print(result)
            print("="*60 + "\n")
            
            return result
    
    except Exception as e:
        print(f"\n✗ SHIPMENT EXTRACTION ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return {'error': f'PDF parsing error: {str(e)}'}
            
# ============================================================================
# API ENDPOINTS
# ============================================================================

@app.route('/')
def index():
    """Serve the main HTML page"""
    return send_file('index.html', mimetype='text/html')

@app.route('/extract-invoice', methods=['POST'])
def extract_invoice():
    """
    Extract invoice data from uploaded PDF
    POST /extract-invoice with file upload
    Returns JSON with extracted fields
    """
    if 'file' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file uploaded'}), 400
    
    file = request.files['file']
    
    if file.filename == '' or not allowed_file(file.filename):
        return jsonify({'status': 'error', 'message': 'Invalid file. Only PDF allowed'}), 400
    
    try:
        # Save uploaded file temporarily
        filename = secure_filename(file.filename)
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)
        
        # Extract data
        extracted_data = extract_invoice_data(filepath)
        
        # Clean up
        os.remove(filepath)
        
        if 'error' in extracted_data:
            return jsonify({'status': 'error', 'message': extracted_data['error']}), 400
        
        return jsonify({
            'status': 'success',
            'data': extracted_data
        })
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/extract-shipment', methods=['POST'])
def extract_shipment():
    """
    Extract shipment data from uploaded PDF
    POST /extract-shipment with file upload
    Returns JSON with extracted fields
    """
    if 'file' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file uploaded'}), 400
    
    file = request.files['file']
    
    if file.filename == '' or not allowed_file(file.filename):
        return jsonify({'status': 'error', 'message': 'Invalid file. Only PDF allowed'}), 400
    
    try:
        # Save uploaded file temporarily
        filename = secure_filename(file.filename)
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)
        
        # Extract data
        extracted_data = extract_shipment_data(filepath)
        
        # Clean up
        os.remove(filepath)
        
        if 'error' in extracted_data:
            return jsonify({'status': 'error', 'message': extracted_data['error']}), 400
        
        return jsonify({
            'status': 'success',
            'data': extracted_data
        })
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/submit-invoice', methods=['POST'])
def submit_invoice():
    """
    Store submitted invoice data in database
    POST /submit-invoice with JSON data
    """
    try:
        data = request.get_json()
        
        # Validate required fields
        if not data.get('invoiceNumber') or not data.get('invoiceDate'):
            return jsonify({'status': 'error', 'message': 'Invoice Number and Date are required'}), 400
        
        conn = sqlite3.connect(DATABASE)
        c = conn.cursor()
        
        c.execute('''
            INSERT INTO invoices 
            (invoice_number, invoice_date, terms_of_payment, terms_of_delivery, currency, amount)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            data.get('invoiceNumber', ''),
            data.get('invoiceDate', ''),
            data.get('termsOfPayment', ''),
            data.get('termsOfDelivery', ''),
            data.get('currency', ''),
            data.get('amount', '')
        ))
        
        conn.commit()
        invoice_id = c.lastrowid
        conn.close()
        
        return jsonify({
            'status': 'success',
            'message': 'Invoice submitted successfully',
            'id': invoice_id
        })
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/submit-shipment', methods=['POST'])
def submit_shipment():
    """
    Store submitted shipment data in database
    POST /submit-shipment with JSON data
    """
    try:
        data = request.get_json()
        
        # Validate required fields
        if not data.get('shipBillNo') or not data.get('shipBillingDate'):
            return jsonify({'status': 'error', 'message': 'Ship Bill No and Date are required'}), 400
        
        conn = sqlite3.connect(DATABASE)
        c = conn.cursor()
        
        c.execute('''
            INSERT INTO shipments 
            (ship_bill_no, ship_billing_date)
            VALUES (?, ?)
        ''', (
            data.get('shipBillNo', ''),
            data.get('shipBillingDate', '')
        ))
        
        conn.commit()
        shipment_id = c.lastrowid
        conn.close()
        
        return jsonify({
            'status': 'success',
            'message': 'Shipment submitted successfully',
            'id': shipment_id
        })
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/export-combined', methods=['GET'])
def export_combined():
    """
    Export all invoices AND shipments to ONE Excel sheet with S.No
    GET /export-combined
    """
    try:
        conn = sqlite3.connect(DATABASE)
        c = conn.cursor()
        
        # Get invoices (without submitted_at)
        c.execute('SELECT invoice_number, invoice_date, terms_of_payment, terms_of_delivery, currency, amount FROM invoices ORDER BY submitted_at DESC')
        invoice_rows = c.fetchall()
        
        # Get shipments (without submitted_at)
        c.execute('SELECT ship_bill_no, ship_billing_date FROM shipments ORDER BY submitted_at DESC')
        shipment_rows = c.fetchall()
        conn.close()
        
        # Create workbook with ONE sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "All Data"
        
        current_row = 1
        
        # ===== INVOICES SECTION =====
        ws[f'A{current_row}'] = "INVOICES"
        ws[f'A{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f'A{current_row}'].fill = PatternFill(start_color="1b5e20", end_color="1b5e20", fill_type="solid")
        current_row += 1
        
        # Invoice headers with S.No (removed Submitted At)
        invoice_headers = ['S.No', 'Invoice Number', 'Invoice Date', 'Terms of Payment', 'Terms of Delivery', 'Currency', 'Amount']
        for col, header in enumerate(invoice_headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="2e7d32", end_color="2e7d32", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1
        
        # Invoice data with S.No
        sno = 1
        for row in invoice_rows:
            ws.cell(row=current_row, column=1).value = sno
            for col, value in enumerate(row, 2):
                ws.cell(row=current_row, column=col).value = value
            sno += 1
            current_row += 1
        
        # Blank row
        current_row += 2
        
        # ===== SHIPMENTS SECTION =====
        ws[f'A{current_row}'] = "SHIPMENTS"
        ws[f'A{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f'A{current_row}'].fill = PatternFill(start_color="1b5e20", end_color="1b5e20", fill_type="solid")
        current_row += 1
        
        # Shipment headers with S.No (removed Submitted At)
        shipment_headers = ['S.No', 'Ship Bill No', 'Ship Billing Date']
        for col, header in enumerate(shipment_headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="2e7d32", end_color="2e7d32", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1
        
        # Shipment data with S.No
        sno = 1
        for row in shipment_rows:
            ws.cell(row=current_row, column=1).value = sno
            for col, value in enumerate(row, 2):
                ws.cell(row=current_row, column=col).value = value
            sno += 1
            current_row += 1
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        return send_file(
            file_stream,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='combined_data.xlsx'
        )
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/get-invoices', methods=['GET'])
def get_invoices():
    """Get all submitted invoices from database"""
    try:
        conn = sqlite3.connect(DATABASE)
        c = conn.cursor()
        c.execute('SELECT id, invoice_number, invoice_date, terms_of_payment, terms_of_delivery, currency, amount, submitted_at FROM invoices ORDER BY submitted_at DESC')
        
        rows = c.fetchall()
        conn.close()
        
        invoices = []
        for row in rows:
            invoices.append({
                'id': row[0],
                'invoiceNumber': row[1],
                'invoiceDate': row[2],
                'termsOfPayment': row[3],
                'termsOfDelivery': row[4],
                'currency': row[5],
                'amount': row[6],
                'submittedAt': row[7]
            })
        
        return jsonify({'status': 'success', 'data': invoices})
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/get-shipments', methods=['GET'])
def get_shipments():
    """Get all submitted shipments from database"""
    try:
        conn = sqlite3.connect(DATABASE)
        c = conn.cursor()
        c.execute('SELECT id, ship_bill_no, ship_billing_date, submitted_at FROM shipments ORDER BY submitted_at DESC')
        
        rows = c.fetchall()
        conn.close()
        
        shipments = []
        for row in rows:
            shipments.append({
                'id': row[0],
                'shipBillNo': row[1],
                'shipBillingDate': row[2],
                'submittedAt': row[3]
            })
        
        return jsonify({'status': 'success', 'data': shipments})
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ============================================================================
# BUNDLE EXTRACTION ENDPOINTS
# ============================================================================

@app.route('/extract-invoice-bundle', methods=['POST'])
def extract_invoice_bundle():
    """
    Extract invoice data from MULTIPLE uploaded PDFs
    POST /extract-invoice-bundle with multiple files
    Returns JSON array with extracted fields from each file
    """
    if 'files' not in request.files:
        return jsonify({'status': 'error', 'message': 'No files uploaded'}), 400
    
    files = request.files.getlist('files')
    
    if not files or len(files) == 0:
        return jsonify({'status': 'error', 'message': 'No files selected'}), 400
    
    try:
        all_extracted_data = []
        
        for file in files:
            if file.filename == '' or not allowed_file(file.filename):
                continue
            
            # Save uploaded file temporarily
            filename = secure_filename(file.filename)
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            file.save(filepath)
            
            # Extract data
            extracted_data = extract_invoice_data(filepath)
            
            # Clean up
            os.remove(filepath)
            
            # Add filename to result
            if 'error' not in extracted_data:
                extracted_data['fileName'] = file.filename
                all_extracted_data.append(extracted_data)
        
        if len(all_extracted_data) == 0:
            return jsonify({'status': 'error', 'message': 'No valid PDF files found'}), 400
        
        return jsonify({
            'status': 'success',
            'data': all_extracted_data,
            'count': len(all_extracted_data)
        })
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/extract-shipment-bundle', methods=['POST'])
def extract_shipment_bundle():
    """
    Extract shipment data from MULTIPLE uploaded PDFs
    POST /extract-shipment-bundle with multiple files
    Returns JSON array with extracted fields from each file
    """
    if 'files' not in request.files:
        return jsonify({'status': 'error', 'message': 'No files uploaded'}), 400
    
    files = request.files.getlist('files')
    
    if not files or len(files) == 0:
        return jsonify({'status': 'error', 'message': 'No files selected'}), 400
    
    try:
        all_extracted_data = []
        
        for file in files:
            if file.filename == '' or not allowed_file(file.filename):
                continue
            
            # Save uploaded file temporarily
            filename = secure_filename(file.filename)
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            file.save(filepath)
            
            # Extract data
            extracted_data = extract_shipment_data(filepath)
            
            # Clean up
            os.remove(filepath)
            
            # Add filename to result
            if 'error' not in extracted_data:
                extracted_data['fileName'] = file.filename
                all_extracted_data.append(extracted_data)
        
        if len(all_extracted_data) == 0:
            return jsonify({'status': 'error', 'message': 'No valid PDF files found'}), 400
        
        return jsonify({
            'status': 'success',
            'data': all_extracted_data,
            'count': len(all_extracted_data)
        })
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    
# ============================================================================
# ERROR HANDLERS
# ============================================================================

@app.errorhandler(404)
def not_found(error):
    return jsonify({'status': 'error', 'message': 'Endpoint not found'}), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({'status': 'error', 'message': 'Internal server error'}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)